"""
Excel API Backend - FastAPI Server
支援多人並發安全的 Excel 檔案操作
Version 3.4.2 - 統一錯誤處理，改善 API 行為一致性
"""

import os
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Depends, Security, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import threading
import time
import logging
from datetime import datetime

load_dotenv()

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Excel API Server",
    description="並發安全的 Excel 檔案操作 API",
    version="3.4.1"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()
API_TOKEN = os.getenv("API_TOKEN", "your-secret-token-here")
EXCEL_ROOT_DIR = Path(os.getenv("EXCEL_ROOT_DIR", "./data"))
EXCEL_ROOT_DIR.mkdir(exist_ok=True)


# ============================================================================
# 文件鎖定管理器
# ============================================================================

class FileLockManager:
    def __init__(self):
        self.locks: Dict[str, threading.Lock] = {}
        self.lock_times: Dict[str, float] = {}
        self._manager_lock = threading.Lock()
        self.default_timeout = float(os.getenv("LOCK_TIMEOUT", "30.0"))
        logger.info(f"FileLockManager initialized with default_timeout={self.default_timeout}s")
    
    def get_lock(self, file_path: str) -> threading.Lock:
        with self._manager_lock:
            if file_path not in self.locks:
                self.locks[file_path] = threading.Lock()
            return self.locks[file_path]
    
    def acquire(self, file_path: str, timeout: float = None) -> bool:
        if timeout is None:
            timeout = self.default_timeout
        
        lock = self.get_lock(file_path)
        start_time = time.time()
        
        while True:
            if lock.acquire(blocking=False):
                self.lock_times[file_path] = time.time()
                logger.info(f"Lock acquired for {file_path}")
                return True
            
            if time.time() - start_time > timeout:
                logger.error(f"Lock timeout for {file_path} after {timeout}s")
                return False
            
            time.sleep(0.1)
    
    def release(self, file_path: str):
        lock = self.get_lock(file_path)
        if lock.locked():
            lock.release()
            elapsed = time.time() - self.lock_times.get(file_path, 0)
            logger.info(f"Lock released for {file_path} (held for {elapsed:.2f}s)")


file_lock_manager = FileLockManager()


# ============================================================================
# Pydantic 模型
# ============================================================================

class AppendRequest(BaseModel):
    file: str = Field(..., description="Excel 檔案名稱")
    sheet: str = Field(default="Sheet1", description="工作表名稱")
    values: List[Any] = Field(..., description="要新增的值列表")

class AppendObjectRequest(BaseModel):
    file: str = Field(..., description="Excel 檔案名稱")
    sheet: str = Field(default="Sheet1", description="工作表名稱")
    values: Dict[str, Any] = Field(..., description="要新增的值(欄位名稱: 值)")

class ReadRequest(BaseModel):
    file: str
    sheet: str = "Sheet1"
    range: Optional[str] = Field(None, description="範圍")

class UpdateAdvancedRequest(BaseModel):
    file: str = Field(..., description="Excel 檔案名稱")
    sheet: str = Field(default="Sheet1", description="工作表名稱")
    row: Optional[int] = Field(None, description="直接指定列號(1-based)")
    lookup_column: Optional[str] = Field(None, description="查找的欄位名稱")
    lookup_value: Optional[str] = Field(None, description="查找的值")
    process_all: Optional[bool] = Field(True, description="是否處理所有匹配記錄(預設True)")
    values_to_set: Dict[str, Any] = Field(..., description="要更新的欄位與值(欄位名稱: 新值)")

class DeleteAdvancedRequest(BaseModel):
    file: str = Field(..., description="Excel 檔案名稱")
    sheet: str = Field(default="Sheet1", description="工作表名稱")
    row: Optional[int] = Field(None, description="直接指定列號(1-based)")
    lookup_column: Optional[str] = Field(None, description="查找的欄位名稱")
    lookup_value: Optional[str] = Field(None, description="查找的值")
    process_all: Optional[bool] = Field(True, description="是否處理所有匹配記錄(預設True)")

class BatchOperation(BaseModel):
    type: str = Field(..., description="操作類型: append, update, delete")
    row: Optional[int] = None
    values: Optional[List[Any]] = None
    column_start: Optional[int] = 1

class BatchRequest(BaseModel):
    file: str
    sheet: str = "Sheet1"
    operations: List[BatchOperation]


# ============================================================================
# 輔助函數
# ============================================================================

def validate_file_path(file_name: str) -> Path:
    if ".." in file_name or "/" in file_name or "\\" in file_name:
        raise HTTPException(status_code=400, detail="Invalid file name")
    file_path = EXCEL_ROOT_DIR / file_name
    return file_path

def get_real_last_row(ws):
    """尋找真正有資料的最後一行"""
    for row_idx in range(ws.max_row, 0, -1):
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col_idx).value not in [None, ""]:
                return row_idx
    return 0

def cleanup_all_empty_rows(ws):
    """徹底清理所有完全空白的行"""
    rows_to_delete = []
    
    for row_idx in range(ws.max_row, 0, -1):
        is_empty = True
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col_idx).value not in [None, ""]:
                is_empty = False
                break
        if is_empty:
            rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete):
        ws.delete_rows(row_idx, 1)
    
    if rows_to_delete:
        logger.info(f"Deleted {len(rows_to_delete)} empty rows")

def cleanup_empty_rows(ws):
    """僅刪除末尾的空白行"""
    last_row = get_real_last_row(ws)
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

def verify_token(credentials: HTTPAuthorizationCredentials = Security(security)):
    if credentials.credentials != API_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    return credentials.credentials

def get_headers(ws) -> Dict[str, int]:
    """
    獲取第一列作為表頭，返回 {欄位名稱: 欄位索引} 的字典
    """
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value:
            headers[str(header_value)] = col_idx
    return headers

def find_all_rows_by_lookup(ws, lookup_column: str, lookup_value: str) -> List[int]:
    """
    根據欄位名稱和值查找所有符合條件的列號
    返回找到的列號列表(1-based)，如果沒找到返回空列表
    """
    headers = get_headers(ws)
    
    if lookup_column not in headers:
        raise HTTPException(
            status_code=400, 
            detail=f"Lookup column '{lookup_column}' not found in headers. Available columns: {list(headers.keys())}"
        )
    
    lookup_col_idx = headers[lookup_column]
    matched_rows = []
    
    # 從第2列開始搜索(第1列是表頭)
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=lookup_col_idx).value
        # 轉換為字串進行比較
        if str(cell_value) == str(lookup_value):
            matched_rows.append(row_idx)
            logger.info(f"Found match at row {row_idx}: {lookup_column}={lookup_value}")
    
    return matched_rows


# ============================================================================
# Excel 操作函數
# ============================================================================

def ensure_file_exists(file_path: Path, sheet_name: str = "Sheet1"):
    if not file_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(file_path)
        logger.info(f"Created new file: {file_path}")

def get_worksheet(file_path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found")
    return wb, wb[sheet_name]

def save_workbook(wb, file_path: Path):
    wb.save(file_path)
    logger.info(f"Saved workbook: {file_path}")


# ============================================================================
# API 端點
# ============================================================================

@app.get("/")
async def root():
    return {
        "service": "Excel API Server",
        "status": "running",
        "version": "3.4.1",
        "timestamp": datetime.now().isoformat(),
        "data_directory": str(EXCEL_ROOT_DIR),
        "lock_timeout": file_lock_manager.default_timeout
    }

@app.get("/api/excel/files")
async def list_files(token: str = Depends(verify_token)):
    try:
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(EXCEL_ROOT_DIR.glob(ext))
        file_list = [f.name for f in excel_files]
        file_list.sort()
        return {"success": True, "files": file_list, "count": len(file_list)}
    except Exception as e:
        logger.error(f"Error listing files: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/sheets")
async def list_sheets(file: str, token: str = Depends(verify_token)):
    file_path = validate_file_path(file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return {"success": True, "sheets": sheet_names}
        finally:
            file_lock_manager.release(str(file_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/headers")
async def get_headers_endpoint(file: str, sheet: str = "Sheet1", token: str = Depends(verify_token)):
    """
    獲取指定工作表的表頭（第一列）
    返回欄位名稱列表，供前端下拉選單使用
    """
    file_path = validate_file_path(file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            if sheet not in wb.sheetnames:
                wb.close()
                raise HTTPException(status_code=404, detail=f"Sheet '{sheet}' not found")
            
            ws = wb[sheet]
            
            # 讀取第一列作為表頭
            headers_dict = get_headers(ws)
            header_names = list(headers_dict.keys())
            
            wb.close()
            
            return {
                "success": True, 
                "headers": header_names,
                "count": len(header_names)
            }
        finally:
            file_lock_manager.release(str(file_path))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting headers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/append")
async def append_row(request: AppendRequest, token: str = Depends(verify_token)):
    """新增一列到 Excel 檔案(陣列模式)"""
    file_path = validate_file_path(request.file)
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            ensure_file_exists(file_path, request.sheet)
            wb, ws = get_worksheet(file_path, request.sheet)
            
            cleanup_all_empty_rows(ws)
            next_row = get_real_last_row(ws) + 1
            
            for col_idx, value in enumerate(request.values, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
            
            save_workbook(wb, file_path)
            return {"success": True, "row_number": next_row}
        finally:
            file_lock_manager.release(str(file_path))
    except Exception as e:
        logger.error(f"Error appending row: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/append_object")
async def append_row_object(request: AppendObjectRequest, token: str = Depends(verify_token)):
    """
    新增一列到 Excel 檔案(物件模式)
    根據欄位名稱自動對應到正確的欄位位置
    """
    file_path = validate_file_path(request.file)
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            ensure_file_exists(file_path, request.sheet)
            wb, ws = get_worksheet(file_path, request.sheet)
            
            # 獲取表頭
            headers = get_headers(ws)
            if not headers:
                raise HTTPException(
                    status_code=400, 
                    detail="No headers found in row 1. Please ensure the first row contains column names."
                )
            
            # 檢查是否有未知的欄位名稱
            unknown_columns = [col for col in request.values.keys() if col not in headers]
            if unknown_columns:
                logger.warning(f"Unknown columns will be ignored: {unknown_columns}")
            
            # 按照表頭順序建立值陣列
            cleanup_all_empty_rows(ws)
            next_row = get_real_last_row(ws) + 1
            
            # 根據表頭順序寫入資料
            for col_name, col_idx in headers.items():
                value = request.values.get(col_name, None)  # 如果沒有提供值，使用 None
                ws.cell(row=next_row, column=col_idx, value=value)
            
            save_workbook(wb, file_path)
            
            return {
                "success": True, 
                "row_number": next_row,
                "matched_columns": [col for col in request.values.keys() if col in headers],
                "ignored_columns": unknown_columns
            }
        finally:
            file_lock_manager.release(str(file_path))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error appending row (object mode): {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/read")
async def read_rows(request: ReadRequest, token: str = Depends(verify_token)):
    file_path = validate_file_path(request.file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if request.sheet not in wb.sheetnames:
                wb.close()
                raise HTTPException(status_code=404, detail=f"Sheet '{request.sheet}' not found")
            ws = wb[request.sheet]
            
            data = []
            rows = ws[request.range] if request.range else ws.rows
            
            for row in rows:
                row_values = []
                for cell in row:
                    val = cell.value
                    if isinstance(val, datetime):
                        fmt = cell.number_format if cell.number_format else ""
                        if any(char in fmt.lower() for char in ['h', 's']) or (':' in fmt):
                            row_values.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            row_values.append(val.strftime('%Y-%m-%d'))
                    else:
                        row_values.append(val)
                
                if any(v not in [None, ""] for v in row_values):
                    data.append(row_values)
            
            return {"success": True, "data": data, "row_count": len(data)}
        finally:
            file_lock_manager.release(str(file_path))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reading file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/excel/update_advanced")
async def update_row_advanced(request: UpdateAdvancedRequest, token: str = Depends(verify_token)):
    """
    進階更新 API - 支持按列號或 Lookup 定位，並按欄位名稱更新
    可選擇處理所有匹配記錄或僅第一筆
    """
    file_path = validate_file_path(request.file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            wb, ws = get_worksheet(file_path, request.sheet)
            
            # 確定要更新的列號
            target_rows = []
            if request.row is not None:
                # 方式1: 直接指定列號
                target_row = request.row
                if target_row < 1 or target_row > ws.max_row:
                    raise HTTPException(status_code=400, detail=f"Invalid row number: {target_row}")
                # 🔒 保護標題列
                if target_row == 1:
                    raise HTTPException(
                        status_code=400, 
                        detail="Cannot update header row (row 1). Data rows start from row 2."
                    )
                target_rows = [target_row]
            elif request.lookup_column and request.lookup_value:
                # 方式2: 透過 Lookup 查找所有符合條件的記錄
                matched_rows = find_all_rows_by_lookup(ws, request.lookup_column, request.lookup_value)
                if not matched_rows:
                    raise HTTPException(
                        status_code=404, 
                        detail=f"No row found where {request.lookup_column} = {request.lookup_value}"
                    )
                # 🆕 根據 process_all 決定處理哪些記錄
                if request.process_all:
                    target_rows = matched_rows  # 處理所有匹配記錄
                else:
                    target_rows = [matched_rows[0]]  # 只處理第一筆
                    logger.info(f"Process mode: First match only (row {matched_rows[0]})")
            else:
                raise HTTPException(
                    status_code=400, 
                    detail="Must provide either 'row' or both 'lookup_column' and 'lookup_value'"
                )
            
            # 獲取表頭
            headers = get_headers(ws)
            
            # 處理單筆或多筆更新
            updated_columns = []
            
            for row_num in target_rows:
                for column_name, new_value in request.values_to_set.items():
                    if column_name not in headers:
                        logger.warning(f"Column '{column_name}' not found in headers, skipping")
                        continue
                    
                    col_idx = headers[column_name]
                    ws.cell(row=row_num, column=col_idx, value=new_value)
                    if column_name not in updated_columns:
                        updated_columns.append(column_name)
                    logger.info(f"Updated row {row_num}, column '{column_name}' = {new_value}")
            
            cleanup_all_empty_rows(ws)
            save_workbook(wb, file_path)
            
            return {
                "success": True, 
                "message": f"{len(target_rows)} row(s) updated",
                "rows_updated": target_rows,
                "updated_count": len(target_rows),
                "updated_columns": updated_columns,
                "process_mode": "all" if request.process_all else "first"
            }
        finally:
            file_lock_manager.release(str(file_path))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating row: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/excel/delete_advanced")
async def delete_row_advanced(request: DeleteAdvancedRequest, token: str = Depends(verify_token)):
    """
    進階刪除 API - 支持按列號或 Lookup 定位
    可選擇處理所有匹配記錄或僅第一筆
    """
    file_path = validate_file_path(request.file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        if not file_lock_manager.acquire(str(file_path)):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            wb, ws = get_worksheet(file_path, request.sheet)
            
            # 確定要刪除的列號
            target_rows = []
            if request.row is not None:
                # 方式1: 直接指定列號
                target_row = request.row
                if target_row < 1 or target_row > ws.max_row:
                    raise HTTPException(status_code=400, detail=f"Invalid row number: {target_row}")
                # 🔒 保護標題列
                if target_row == 1:
                    raise HTTPException(
                        status_code=400, 
                        detail="Cannot delete header row (row 1). Data rows start from row 2."
                    )
                target_rows = [target_row]
            elif request.lookup_column and request.lookup_value:
                # 方式2: 透過 Lookup 查找所有符合條件的記錄
                matched_rows = find_all_rows_by_lookup(ws, request.lookup_column, request.lookup_value)
                if not matched_rows:
                    raise HTTPException(
                        status_code=404, 
                        detail=f"No row found where {request.lookup_column} = {request.lookup_value}"
                    )
                # 🆕 根據 process_all 決定處理哪些記錄
                if request.process_all:
                    target_rows = matched_rows  # 處理所有匹配記錄
                else:
                    target_rows = [matched_rows[0]]  # 只處理第一筆
                    logger.info(f"Process mode: First match only (row {matched_rows[0]})")
            else:
                raise HTTPException(
                    status_code=400, 
                    detail="Must provide either 'row' or both 'lookup_column' and 'lookup_value'"
                )
            
            # 處理單筆或多筆刪除(從後往前刪除以避免行號偏移)
            rows_to_delete = sorted(target_rows, reverse=True)
            
            for row_num in rows_to_delete:
                ws.delete_rows(row_num)
                logger.info(f"Deleted row {row_num}")
            
            cleanup_all_empty_rows(ws)
            save_workbook(wb, file_path)
            
            return {
                "success": True, 
                "message": f"{len(target_rows)} row(s) deleted",
                "rows_deleted": target_rows,
                "deleted_count": len(target_rows),
                "process_mode": "all" if request.process_all else "first"
            }
        finally:
            file_lock_manager.release(str(file_path))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting row: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/batch")
async def batch_operations(request: BatchRequest, token: str = Depends(verify_token)):
    file_path = validate_file_path(request.file)
    try:
        if not file_lock_manager.acquire(str(file_path), timeout=60.0):
            raise HTTPException(status_code=503, detail="File is locked")
        try:
            ensure_file_exists(file_path, request.sheet)
            wb, ws = get_worksheet(file_path, request.sheet)
            
            cleanup_all_empty_rows(ws)
            
            results = []
            for op in request.operations:
                try:
                    if op.type == "append":
                        nr = get_real_last_row(ws) + 1
                        for ci, v in enumerate(op.values, 1):
                            ws.cell(row=nr, column=ci, value=v)
                        results.append({"operation": "append", "success": True, "row_number": nr})
                    elif op.type == "update":
                        for ci, v in enumerate(op.values, op.column_start):
                            ws.cell(row=op.row, column=ci, value=v)
                        results.append({"operation": "update", "success": True, "row": op.row})
                    elif op.type == "delete":
                        ws.delete_rows(op.row)
                        results.append({"operation": "delete", "success": True, "row": op.row})
                except Exception as e:
                    results.append({"operation": op.type, "success": False, "error": str(e)})
            
            save_workbook(wb, file_path)
            return {"success": True, "results": results}
        finally:
            file_lock_manager.release(str(file_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8000"))
    
    logger.info(f"Starting Excel API Server v3.4.2")
    logger.info(f"Server address: {host}:{port}")
    uvicorn.run(app, host=host, port=port)
